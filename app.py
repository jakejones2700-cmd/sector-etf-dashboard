import streamlit as st
import yfinance as yf
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, timedelta, date
import numpy as np
import io

st.set_page_config(
    page_title="S&P Sector ETF Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Custom CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main { background-color: #0e1117; }
    .metric-card {
        background: #1c2333;
        border-radius: 10px;
        padding: 14px 16px;
        margin: 4px 0;
        border-left: 4px solid #444;
    }
    .metric-card.positive { border-left-color: #00c853; }
    .metric-card.negative { border-left-color: #ff1744; }
    .metric-card.neutral  { border-left-color: #888; }
    .etf-name  { font-size: 13px; color: #aaa; margin-bottom: 2px; }
    .etf-ticker{ font-size: 18px; font-weight: 700; color: #fff; }
    .etf-ret   { font-size: 20px; font-weight: 700; }
    .etf-price { font-size: 13px; color: #aaa; }
    .positive-text { color: #00c853; }
    .negative-text { color: #ff1744; }
    h1 { color: #ffffff !important; }
    .stSelectbox label, .stDateInput label { color: #ccc !important; }
    div[data-testid="stMetricValue"] { font-size: 20px !important; }
</style>
""", unsafe_allow_html=True)

# ── Sector ETF Universe ─────────────────────────────────────────────────────
SECTORS = {
    "XLK":  "Technology",
    "XLF":  "Financials",
    "XLV":  "Health Care",
    "XLY":  "Consumer Disc.",
    "XLP":  "Consumer Staples",
    "XLE":  "Energy",
    "XLI":  "Industrials",
    "XLB":  "Materials",
    "XLRE": "Real Estate",
    "XLU":  "Utilities",
    "XLC":  "Communication Svcs",
    "SPY":  "S&P 500 (Benchmark)",
}

SECTOR_COLORS = {
    "XLK":  "#4fc3f7",
    "XLF":  "#81c784",
    "XLV":  "#f48fb1",
    "XLY":  "#ffb74d",
    "XLP":  "#ce93d8",
    "XLE":  "#ffcc02",
    "XLI":  "#80cbc4",
    "XLB":  "#a5d6a7",
    "XLRE": "#ef9a9a",
    "XLU":  "#b0bec5",
    "XLC":  "#90caf9",
    "SPY":  "#ffffff",
}

# ── Time Period Helpers ─────────────────────────────────────────────────────
def last_trading_close_before(d: date) -> date:
    """Return the last calendar day before `d` that could be a trading day.
    We fetch a small window and yfinance will return the last available close."""
    return d - timedelta(days=1)

def get_period_dates(period: str):
    today = date.today()
    if period == "1D":
        # Go back enough to guarantee we get the prior close
        start = today - timedelta(days=5)
    elif period == "1W":
        start = today - timedelta(weeks=1)
    elif period == "1M":
        start = today - timedelta(days=30)
    elif period == "3M":
        start = today - timedelta(days=91)
    elif period == "6M":
        start = today - timedelta(days=182)
    elif period == "MTD":
        # Anchor to last close of prior month (day before the 1st)
        first_of_month = today.replace(day=1)
        start = last_trading_close_before(first_of_month)
    elif period == "QTD":
        # Anchor to last close of prior quarter
        q_start_month = ((today.month - 1) // 3) * 3 + 1
        first_of_quarter = today.replace(month=q_start_month, day=1)
        start = last_trading_close_before(first_of_quarter)
    elif period == "YTD":
        # Anchor to last close of prior year (Dec 31)
        start = last_trading_close_before(today.replace(month=1, day=1))
    elif period == "1Y":
        start = today - timedelta(days=365)
    elif period == "3Y":
        start = today - timedelta(days=365 * 3)
    elif period == "5Y":
        start = today - timedelta(days=365 * 5)
    elif period == "ALL":
        start = date(1998, 1, 1)
    else:
        start = today - timedelta(days=30)
    return start, today

@st.cache_data(ttl=300)
def fetch_data(tickers: list, start: date, end: date) -> pd.DataFrame:
    raw = yf.download(
        tickers,
        start=start.strftime("%Y-%m-%d"),
        end=(end + timedelta(days=1)).strftime("%Y-%m-%d"),
        auto_adjust=True,
        progress=False,
        threads=True,
    )
    if isinstance(raw.columns, pd.MultiIndex):
        closes = raw["Close"]
    else:
        closes = raw[["Close"]]
        closes.columns = tickers
    closes = closes.dropna(how="all")
    return closes

def calc_return(series: pd.Series, anchor_last_of_start: bool = False) -> float | None:
    """
    anchor_last_of_start=True: used for MTD/QTD/YTD where we fetched a small
    window ending on the last trading day before the period. We use that window's
    LAST close as the base, and the series' overall last close as the end.
    For normal periods, just use first vs last.
    """
    s = series.dropna()
    if len(s) < 2:
        return None
    return (s.iloc[-1] / s.iloc[0] - 1) * 100

def fmt_ret(val):
    if val is None:
        return "N/A", "neutral"
    sign = "+" if val >= 0 else ""
    css = "positive" if val >= 0 else "negative"
    return f"{sign}{val:.2f}%", css

# ── Sidebar ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Settings")

    period_options = ["1D","1W","1M","3M","6M","MTD","QTD","YTD","1Y","3Y","5Y","ALL","Custom"]
    selected_period = st.selectbox("Time Period", period_options, index=6)

    custom_start = custom_end = None
    if selected_period == "Custom":
        custom_start = st.date_input("Start Date", value=date.today() - timedelta(days=90))
        custom_end   = st.date_input("End Date",   value=date.today())

    st.markdown("---")
    chart_type = st.selectbox("Chart Type", ["Normalized Return (%)", "Price Level"])
    show_benchmark = st.checkbox("Show SPY Benchmark on Chart", value=True)

    st.markdown("---")
    selected_etfs = st.multiselect(
        "ETFs to Display",
        options=list(SECTORS.keys()),
        default=list(SECTORS.keys()),
        format_func=lambda x: f"{x} – {SECTORS[x]}"
    )

    st.markdown("---")
    st.caption("Data via Yahoo Finance · Refreshes every 5 min")
    if st.button("🔄 Refresh Data"):
        st.cache_data.clear()
        st.rerun()

# ── Main ────────────────────────────────────────────────────────────────────
st.markdown("# 📊 S&P Sector ETF Performance Dashboard")

if not selected_etfs:
    st.warning("Select at least one ETF in the sidebar.")
    st.stop()

# Resolve dates
if selected_period == "Custom" and custom_start and custom_end:
    start_date, end_date = custom_start, custom_end
else:
    start_date, end_date = get_period_dates(selected_period)

# Always fetch SPY for benchmark even if deselected
all_tickers = list(set(selected_etfs + ["SPY"]))

with st.spinner("Fetching latest prices…"):
    closes = fetch_data(all_tickers, start_date, end_date)

if closes.empty:
    st.error("No data returned. Try a different date range.")
    st.stop()

# ── KPI Row ─────────────────────────────────────────────────────────────────
st.markdown(f"### Performance — {selected_period}  &nbsp; `{start_date}` → `{end_date}`")

# Sort by return descending
returns = {}
for tk in selected_etfs:
    if tk in closes.columns:
        returns[tk] = calc_return(closes[tk])

sorted_tickers = sorted(returns.keys(), key=lambda x: (returns[x] is None, -(returns[x] or 0)))

cols = st.columns(min(len(sorted_tickers), 6))
for i, tk in enumerate(sorted_tickers):
    ret = returns[tk]
    ret_str, css = fmt_ret(ret)
    price = closes[tk].dropna().iloc[-1] if tk in closes.columns and not closes[tk].dropna().empty else None
    price_str = f"${price:.2f}" if price else "N/A"
    color = "#00c853" if css == "positive" else ("#ff1744" if css == "negative" else "#aaa")
    with cols[i % 6]:
        st.markdown(f"""
        <div class="metric-card {css}">
            <div class="etf-ticker">{tk}</div>
            <div class="etf-name">{SECTORS.get(tk,'')}</div>
            <div class="etf-ret" style="color:{color}">{ret_str}</div>
            <div class="etf-price">{price_str}</div>
        </div>""", unsafe_allow_html=True)

st.markdown("---")

# ── Main Chart ──────────────────────────────────────────────────────────────
st.markdown("### 📈 Performance Chart")

fig = go.Figure()
chart_tickers = selected_etfs if not show_benchmark else list(set(selected_etfs + ["SPY"]))

for tk in chart_tickers:
    if tk not in closes.columns:
        continue
    s = closes[tk].dropna()
    if s.empty:
        continue
    if chart_type == "Normalized Return (%)":
        y = (s / s.iloc[0] - 1) * 100
        yaxis_label = "Return (%)"
    else:
        y = s
        yaxis_label = "Price (USD)"

    is_spy = tk == "SPY"
    fig.add_trace(go.Scatter(
        x=s.index,
        y=y,
        name=f"{tk} – {SECTORS.get(tk,'')}",
        line=dict(
            color=SECTOR_COLORS.get(tk, "#888"),
            width=2.5 if not is_spy else 1.5,
            dash="dot" if is_spy and tk not in selected_etfs else "solid",
        ),
        opacity=0.6 if is_spy and tk not in selected_etfs else 1.0,
        hovertemplate=f"<b>{tk}</b><br>%{{x|%b %d, %Y}}<br>{'Return' if chart_type=='Normalized Return (%)' else 'Price'}: %{{y:.2f}}{'%' if chart_type=='Normalized Return (%)' else ' USD'}<extra></extra>"
    ))

fig.update_layout(
    template="plotly_dark",
    paper_bgcolor="#0e1117",
    plot_bgcolor="#0e1117",
    hovermode="x unified",
    legend=dict(orientation="v", x=1.01, y=1, bgcolor="rgba(0,0,0,0)"),
    xaxis=dict(showgrid=False, zeroline=False),
    yaxis=dict(showgrid=True, gridcolor="#1e2a3a", zeroline=True, zerolinecolor="#444", title=yaxis_label),
    margin=dict(l=0, r=0, t=20, b=0),
    height=480,
)

if chart_type == "Normalized Return (%)" and selected_period not in ["ALL","5Y","3Y"]:
    fig.add_hline(y=0, line_dash="dash", line_color="#555", line_width=1)

st.plotly_chart(fig, use_container_width=True)

# ── Bar Chart: Ranking ───────────────────────────────────────────────────────
st.markdown("### 🏆 Sector Return Ranking")
bar_data = [(tk, returns[tk]) for tk in sorted_tickers if returns[tk] is not None]
if bar_data:
    tks, rets = zip(*bar_data)
    bar_colors = ["#00c853" if r >= 0 else "#ff1744" for r in rets]
    fig2 = go.Figure(go.Bar(
        x=list(tks),
        y=list(rets),
        marker_color=bar_colors,
        text=[f"{r:+.2f}%" for r in rets],
        textposition="outside",
        hovertemplate="<b>%{x}</b><br>Return: %{y:.2f}%<extra></extra>",
    ))
    fig2.update_layout(
        template="plotly_dark",
        paper_bgcolor="#0e1117",
        plot_bgcolor="#0e1117",
        yaxis=dict(showgrid=True, gridcolor="#1e2a3a", zeroline=True, zerolinecolor="#555", title="Return (%)"),
        xaxis=dict(showgrid=False),
        margin=dict(l=0, r=0, t=30, b=0),
        height=350,
        showlegend=False,
    )
    st.plotly_chart(fig2, use_container_width=True)

# ── Multi-Period Summary Table ──────────────────────────────────────────────
st.markdown("### 📋 Multi-Period Return Summary")

SUMMARY_PERIODS = ["1D","1W","MTD","1M","QTD","3M","YTD","6M","1Y","3Y","5Y"]

@st.cache_data(ttl=300)
def build_summary(tickers):
    rows = []
    for period in SUMMARY_PERIODS:
        s, e = get_period_dates(period)
        data = fetch_data(tickers, s, e)
        row = {"Period": period}
        for tk in tickers:
            if tk in data.columns:
                row[tk] = calc_return(data[tk])
            else:
                row[tk] = None
        rows.append(row)
    return pd.DataFrame(rows).set_index("Period")

with st.spinner("Building summary table…"):
    summary_df = build_summary(tuple(selected_etfs))

def color_val(val):
    if pd.isna(val) or val is None:
        return "color: #555"
    return "color: #00c853" if val >= 0 else "color: #ff1744"

def fmt_cell(val):
    if pd.isna(val) or val is None:
        return "N/A"
    return f"{'+' if val >= 0 else ''}{val:.2f}%"

try:
    styled = summary_df.style.map(color_val).format(fmt_cell)
except AttributeError:
    styled = summary_df.style.applymap(color_val).format(fmt_cell)
st.dataframe(styled, use_container_width=True, height=430)


# ── Excel Download ─────────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("### 📥 Download Data")

def build_excel(closes_df, tickers):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # Sheet 1: Daily Prices
        price_df = closes_df[tickers].copy()
        price_df.index = pd.to_datetime(price_df.index).strftime("%Y-%m-%d")
        price_df.index.name = "Date"
        price_df.columns = [f"{tk} Price" for tk in tickers]
        price_df.to_excel(writer, sheet_name="Daily Prices")

        # Sheet 2: Daily $ Change
        dollar_chg = closes_df[tickers].diff()
        dollar_chg.index = pd.to_datetime(dollar_chg.index).strftime("%Y-%m-%d")
        dollar_chg.index.name = "Date"
        dollar_chg.columns = [f"{tk} $ Chg" for tk in tickers]
        dollar_chg.to_excel(writer, sheet_name="Daily $ Change")

        # Sheet 3: Daily % Change
        pct_chg = closes_df[tickers].pct_change() * 100
        pct_chg.index = pd.to_datetime(pct_chg.index).strftime("%Y-%m-%d")
        pct_chg.index.name = "Date"
        pct_chg.columns = [f"{tk} % Chg" for tk in tickers]
        pct_chg.to_excel(writer, sheet_name="Daily % Change")

        # Sheet 4: Combined long format
        frames = []
        for tk in tickers:
            s = closes_df[tk].dropna()
            frames.append(pd.DataFrame({
                "Date": pd.to_datetime(s.index).strftime("%Y-%m-%d"),
                "Ticker": tk,
                "Sector": SECTORS.get(tk, ""),
                "Close Price": s.values,
                "$ Change": s.diff().values,
                "% Change": s.pct_change().mul(100).values,
                "Cumulative Return (%)": (s / s.iloc[0] - 1).mul(100).values,
            }))
        pd.concat(frames, ignore_index=True).to_excel(writer, sheet_name="Combined", index=False)

        # Header formatting
        from openpyxl.styles import Font, PatternFill, Alignment
        hfont = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="1F3864")
        for sname in writer.sheets:
            ws = writer.sheets[sname]
            for cell in ws[1]:
                cell.font = hfont
                cell.fill = hfill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                w = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = w + 4

    output.seek(0)
    return output.read()

col_dl1, col_dl2 = st.columns([2, 3])
with col_dl1:
    dl_tickers = st.multiselect(
        "ETFs to include in export",
        options=selected_etfs,
        default=selected_etfs,
        key="dl_tickers"
    )
with col_dl2:
    st.markdown("<br>", unsafe_allow_html=True)
    if dl_tickers:
        valid = [tk for tk in dl_tickers if tk in closes.columns]
        if valid:
            export_df = closes[valid].dropna(how='all')
            excel_bytes = build_excel(export_df, valid)
            fname = f"sector_etf_{selected_period}_{date.today().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label="⬇️ Download Excel (.xlsx)",
                data=excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    else:
        st.info("Select at least one ETF above to enable download.")

st.caption("File includes 4 sheets: Daily Prices · Daily $ Change · Daily % Change · Combined")

# ── Footer ───────────────────────────────────────────────────────────────────
st.markdown("---")
st.caption("📌 Data sourced from Yahoo Finance via `yfinance`. Prices are delayed ~15 min during market hours. Not financial advice.")
